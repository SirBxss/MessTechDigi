import os
import win32com.client as win32
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import pandas as pd
from typing import Optional, List


def delete_rows_with_shapes(file_path: str, sheet_name: str, temp_file_path: str) -> None:
    """
    Delete rows in an Excel sheet that contain shapes.
    """
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        workbook = excel.Workbooks.Open(file_path)
        sheet = workbook.Sheets(sheet_name)

        shape_rows = set()
        # Identify rows that contain shapes
        for shape in sheet.Shapes:
            top_row = shape.TopLeftCell.Row
            bottom_row = shape.BottomRightCell.Row
            shape_rows.update(range(top_row, bottom_row + 1))

        # Delete rows that contain shapes
        for row in sorted(shape_rows, reverse=True):
            sheet.Rows(row).Delete()

        workbook.SaveAs(temp_file_path)
        workbook.Close(SaveChanges=False)
        excel.Quit()
    except Exception as e:
        print(f"Error in delete_rows_with_shapes: {e}")


def clean_and_format_excel(file_path: str, save_path: str) -> pd.DataFrame:
    """
    Clean and format an Excel file:
    - Handle merged cells.
    - Identify 'Date of Birth' and 'ID' columns.
    - Format date cells.
    - Delete rows with empty IDs.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        handle_merged_cells(sheet)
        header_row = 1
        dob_column_index, id_column_index = identify_headers(sheet, header_row)
        rows_to_delete = format_dates_and_identify_empty_ids(sheet, dob_column_index, id_column_index)

        # Delete rows with empty IDs
        for row in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row)

        df = convert_sheet_to_dataframe(sheet)
        df.to_excel(save_path, index=False)

        return df
    except Exception as e:
        print(f"Error in clean_and_format_excel: {e}")
        return pd.DataFrame()


def handle_merged_cells(sheet: Worksheet) -> None:
    """
    Unmerge cells and fill them with the top-left value.
    """
    merged_ranges = list(sheet.merged_cells.ranges)
    # Unmerge cells and fill with top-left value
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = top_left_value


def identify_headers(sheet: Worksheet, header_row: int) -> (Optional[int], Optional[int]):
    """
    Identify the 'Date of Birth' and 'ID' columns based on common headers.
    """
    dob_column_index = None
    id_column_index = None
    # Identify columns based on header values
    for col in range(1, sheet.max_column + 1):
        header_value = str(sheet.cell(row=header_row, column=col).value).strip().lower()
        if header_value in ['date of birth', 'dob', 'birthdate', 'birth date', 'geburtsdatum']:
            dob_column_index = col
        if header_value in ['id', 'identification number', 'identifier', 'identification']:
            id_column_index = col

    if dob_column_index is None:
        raise ValueError("Date of Birth column not found")
    if id_column_index is None:
        raise ValueError("ID column not found")

    return dob_column_index, id_column_index


def format_dates_and_identify_empty_ids(sheet: Worksheet, dob_col_idx: int, id_col_idx: int) -> List[int]:
    """
    Format date cells and identify rows with empty IDs.
    """
    rows_to_delete = []
    # Format date cells and mark rows with empty IDs
    for row in range(2, sheet.max_row + 1):
        id_cell = sheet.cell(row=row, column=id_col_idx)
        dob_cell = sheet.cell(row=row, column=dob_col_idx)
        if not id_cell.value or str(id_cell.value).strip() == "":
            rows_to_delete.append(row)
        if isinstance(dob_cell.value, datetime):
            dob_cell.value = dob_cell.value.strftime('%d.%m.%Y')
    return rows_to_delete


def convert_sheet_to_dataframe(sheet: Worksheet) -> pd.DataFrame:
    """
    Convert an Excel sheet to a pandas DataFrame.
    """
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]  # Set first row as header
    df = df[1:]  # Remove the first row
    return df


def search_id_and_retrieve_rows(df: pd.DataFrame, search_id: str) -> None:
    """
    Search for rows matching the given ID and print them.
    """
    filtered_df = df[df['ID'] == search_id]
    print(filtered_df.to_string(index=False))


if __name__ == "__main__":
    # Receive inputs from the user
    original_file_path = input("Enter the path to the original Excel file: ")
    directory = os.path.dirname(original_file_path)
    temp_file_path = directory + '\Intermediate.xlsx'
    cleaned_file_path = directory + '\Cleaned_sample_list.xlsx'
    sheet_name = input("Enter the sheet name: ")
    search_id = input("Enter the ID to search for: ")

    delete_rows_with_shapes(original_file_path, sheet_name, temp_file_path)
    cleaned_df = clean_and_format_excel(temp_file_path, cleaned_file_path)
    search_id_and_retrieve_rows(cleaned_df, search_id)
